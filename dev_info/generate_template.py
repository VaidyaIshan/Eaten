"""
Generate example Excel template for user seeding.
Run this script to create an example Excel file for user seeding.

Usage: python generate_template.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_template():
    # Create workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    example_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Create headers
    headers = ['username', 'email', 'password', 'role_id', 'is_active']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add example data
    examples = [
        ('john_doe', 'john@example.com', 'securepass123', 2, 'true'),
        ('jane_smith', 'jane@example.com', 'password456', 1, 'true'),
        ('bob_wilson', 'bob@example.com', 'mypassword789', 2, 'false'),
        ('alice_johnson', 'alice@example.com', 'alicepass999', 2, 'true'),
    ]
    
    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.fill = example_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add empty rows for user input (with borders)
    for row_num in range(6, 16):  # 10 empty rows
        for col_num in range(1, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Add instructions sheet
    instructions_sheet = wb.create_sheet("Instructions")
    
    instructions_sheet.column_dimensions['A'].width = 80
    
    instructions = [
        ("USER SEEDING INSTRUCTIONS", None),
        ("", None),
        ("Column Descriptions:", None),
        ("", None),
        ("1. username (Required)", "Unique identifier for the user. No spaces or special characters recommended."),
        ("2. email (Required)", "Valid email address. Must be unique in the system."),
        ("3. password (Required)", "Password for the user. Will be hashed before storing."),
        ("4. role_id (Optional)", "User role: 0=Superadmin, 1=Admin, 2=User (default). Defaults to 2 if empty."),
        ("5. is_active (Optional)", "User status: true/false or 1/0. Defaults to false if empty."),
        ("", None),
        ("IMPORTANT NOTES:", None),
        ("", None),
        ("• Do NOT modify the column headers", None),
        ("• All entries in username and email columns must be UNIQUE", None),
        ("• Use the 'Users' sheet to add your data", None),
        ("• You can delete the example rows and add your own", None),
        ("• The system will validate each row individually", None),
        ("• If a row has errors, it will be skipped and reported", None),
        ("• Other rows will continue processing normally", None),
        ("• For role_id: Use 2 for regular users, 1 for admins", None),
        ("• Superadmin (0) should only be used for special accounts", None),
        ("", None),
        ("EXAMPLE VALID DATA:", None),
        ("", None),
        ("Username: john_doe, Email: john@example.com, Password: secure123, Role: 2, Active: true", None),
        ("Username: admin_user, Email: admin@example.com, Password: admin456, Role: 1, Active: true", None),
        ("", None),
        ("ROLE IDs REFERENCE:", None),
        ("0 = Superadmin (Full system access)", None),
        ("1 = Admin (Admin panel access)", None),
        ("2 = User (Regular user, default)", None),
    ]
    
    for row_num, (title, description) in enumerate(instructions, 1):
        cell_title = instructions_sheet.cell(row=row_num, column=1)
        cell_title.value = title
        cell_title.font = Font(bold=True, size=11) if description is None else Font(bold=True, size=10, color="1F4E78")
        cell_title.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        if description:
            cell_desc = instructions_sheet.cell(row=row_num, column=2)
            cell_desc.value = description
            cell_desc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell_desc.font = Font(size=10)
    
    # Save the file
    filename = f"user_seeding_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✓ Template created successfully: {filename}")
    print(f"")
    print(f"Template includes:")
    print(f"  • 'Users' sheet with column headers and example data")
    print(f"  • 'Instructions' sheet with detailed guidelines")
    print(f"  • Pre-formatted columns and example rows")
    print(f"  • 10 empty rows ready for your data")

if __name__ == "__main__":
    create_template()
